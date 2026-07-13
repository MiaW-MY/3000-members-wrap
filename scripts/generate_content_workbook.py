#!/usr/bin/env python3
"""Generate BA Career 3000 Wrap content workbook for ops team."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "docs/BA_Career_3000_Wrap_Content.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="6B4E9B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="6B4E9B")
NOTE_FONT = Font(italic=True, color="666666", size=10)
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
UPDATE_FILL = PatternFill("solid", fgColor="FFF9E6")


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_data_rows(ws, start_row, end_row, cols, update_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if c == update_col:
                cell.fill = UPDATE_FILL


def write_sheet(ws, title, rows, note=None):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    start = 3
    if note:
        ws["A2"] = note
        ws["A2"].font = NOTE_FONT
        ws.merge_cells("A2:D2")
        start = 4

    headers = ["Type", "Field", "Current Content", "Update"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)

    style_header_row(ws, start, 4)

    for idx, row in enumerate(rows, start + 1):
        for col, val in enumerate(row, 1):
            ws.cell(row=idx, column=col, value=val)

    style_data_rows(ws, start + 1, start + len(rows), 4, 4)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 58
    ws.column_dimensions["D"].width = 40
    ws.freeze_panes = ws.cell(row=start + 1, column=1).coordinate


def build_workbook():
    wb = Workbook()

    # --- Overview ---
    ov = wb.active
    ov.title = "Overview"
    ov["A1"] = "BA Career Meetup — 3,000 Wrap · Page Overview"
    ov["A1"].font = TITLE_FONT
    ov.merge_cells("A1:E1")
    ov["A2"] = "Fill in the Update column on each page sheet when requesting changes. Keep anchor numbers consistent (see Consistency Rules sheet)."
    ov["A2"].font = NOTE_FONT
    ov.merge_cells("A2:E2")

    ov_headers = ["Page #", "Page Name", "Screen ID", "Swipe Order", "Notes"]
    for i, h in enumerate(ov_headers, 1):
        ov.cell(row=4, column=i, value=h)
    style_header_row(ov, 4, 5)

    overview_rows = [
        ("01", "Opening", "opening", "1", "Full-screen hero; no top bar logo"),
        ("02", "Our Journey", "journey-intro", "2", "Vertical timeline 2023–2026"),
        ("03", "By the Numbers", "stats", "3", "Six stat cards"),
        ("04", "Community Moments", "moments", "4", "Photo collage (5 tiles)"),
        ("05", "Thank You — Speakers", "thank-speakers", "5", ""),
        ("06", "Thank You — Partners", "thank-partners", "6", "Up to 12 logos shown on page"),
        ("07", "Thank You — Volunteers", "thank-volunteers", "7", ""),
        ("08", "What's Next", "whats-next", "8", "Three cards + bottom group photo"),
        ("09", "Your Voice", "feedback", "9", "Feedback form"),
        ("10", "Closing", "closing", "10", "Full-page artwork image"),
    ]
    for idx, row in enumerate(overview_rows, 5):
        for col, val in enumerate(row, 1):
            ov.cell(row=idx, column=col, value=val)
    style_data_rows(ov, 5, 5 + len(overview_rows) - 1, 5, 5)

    ov.column_dimensions["A"].width = 10
    ov.column_dimensions["B"].width = 26
    ov.column_dimensions["C"].width = 18
    ov.column_dimensions["D"].width = 14
    ov.column_dimensions["E"].width = 36
    ov.freeze_panes = "A5"

    # --- Page sheets ---
    pages = {
        "01 Opening": (
            "01 · Opening",
            "Replace collage photos with same filenames, or note new filenames in Update.",
            [
                ("Copy", "Badge", "WE DID IT TOGETHER!", ""),
                ("Copy", "Headline number", "3,000", ""),
                ("Copy", "Headline word", "Members", ""),
                ("Copy", "Subcopy line 1", "One growing community. Thank you", ""),
                ("Copy", "Subcopy line 2", "for being part of our journey.", ""),
                ("Copy", "Footer hint", "Swipe up to explore", ""),
                ("Image", "Logo", "assets/icons/Icon black.png", ""),
                ("Image", "Collage photo 1", "assets/opening/collage-1.jpg", ""),
                ("Image", "Collage photo 2", "assets/opening/collage-2.jpg", ""),
                ("Image", "Collage photo 3", "assets/opening/collage-3.jpg", ""),
                ("Image", "Collage photo 4", "assets/opening/collage-4.jpg", ""),
                ("Image", "Collage photo 5", "assets/opening/collage-5.jpg", ""),
                ("Image", "Collage photo 6", "assets/opening/collage-6.jpg", ""),
                ("Interaction", "Play button", "Decorative only — no video linked yet", ""),
            ],
        ),
        "02 Our Journey": (
            "02 · Our Journey",
            "Keep 3,000 and 2,000 milestones aligned with Stats and Closing pages.",
            [
                ("Copy", "Title", "Our Journey", ""),
                ("Copy", "Subtitle", "From a small idea in 2023 to 3,000 members today.", ""),
                ("Copy", "2023 timeline text", "We started with a simple idea, to support Business Analysts and build a welcoming space.", ""),
                ("Copy", "2024 timeline text", "More events, more connections, stronger community across Aotearoa New Zealand.", ""),
                ("Copy", "2025 timeline text", "We celebrated 2,000 members and saw the community continue to grow through events, stories and shared support.", ""),
                ("Copy", "2026 timeline text (highlight)", "Now we're 3,000 strong, and the journey is still growing.", ""),
                ("Copy", "Swipe hint", "Swipe up to next", ""),
            ],
        ),
        "03 By the Numbers": (
            "03 · By the Numbers",
            "Verify all numbers against source data before updating.",
            [
                ("Copy", "Title", "By the Numbers", ""),
                ("Copy", "Subtitle", "A look back at our impact together.", ""),
                ("Data", "Members", "3,000+", ""),
                ("Data", "Events Hosted", "40+ · Online & in-person", ""),
                ("Data", "Guest Speakers", "60+ · Sharing their stories", ""),
                ("Data", "Volunteers", "22+ · Powering our community", ""),
                ("Data", "Countries", "35+ · Our members come from", ""),
                ("Data", "Avg. Event Rating", "4.8 / 5 · From 341 feedbacks", ""),
                ("Copy", "Footer note", "Thank you for making these numbers meaningful.", ""),
            ],
        ),
        "04 Community Moments": (
            "04 · Community Moments",
            "Five landscape event photos recommended.",
            [
                ("Copy", "Title line 1", "Community", ""),
                ("Copy", "Title line 2", "Moments", ""),
                ("Copy", "Subtitle", "Behind every milestone are the people, stories and conversations that brought us here.", ""),
                ("Image", "Photo 1", "assets/moments/moment-1.jpg", ""),
                ("Copy", "Photo 1 caption (accessibility)", "Learning together", ""),
                ("Image", "Photo 2", "assets/moments/moment-2.jpg", ""),
                ("Copy", "Photo 2 caption (accessibility)", "Networking night", ""),
                ("Image", "Photo 3", "assets/moments/moment-3.jpg", ""),
                ("Copy", "Photo 3 caption (accessibility)", "Sharing stories", ""),
                ("Image", "Photo 4", "assets/moments/moment-4.jpg", ""),
                ("Copy", "Photo 4 caption (accessibility)", "Growing together", ""),
                ("Image", "Photo 5", "assets/moments/moment-5.jpg", ""),
                ("Copy", "Photo 5 caption (accessibility)", "Community in action", ""),
            ],
        ),
        "05 Thank You Speakers": (
            "05 · Thank You — Speakers",
            None,
            [
                ("Copy", "Title", "Our Speakers", ""),
                ("Copy", "Intro", "Thank you for sharing your stories, experience and insights. You helped our community learn, reflect and grow.", ""),
                ("Copy", "Featured badge", "Featured Talk", ""),
                ("Copy", "Quote 1", "Every talk adds a new perspective.", ""),
                ("Copy", "Quote 2", "Ideas shared today, impact tomorrow.", ""),
                ("Image", "Title decoration", "assets/decor/speakers-heart-transparent.png", ""),
                ("Image", "Main photo", "assets/thanks/speakers-1.jpg", ""),
                ("Image", "Photo 2", "assets/thanks/speakers-2.jpg", ""),
                ("Image", "Photo 3", "assets/thanks/speakers-3.jpg", ""),
            ],
        ),
        "06 Thank You Partners": (
            "06 · Thank You — Partners",
            "Partner names must match official branding. Page currently displays up to 12 logos.",
            [
                ("Copy", "Title", "Our Partners", ""),
                ("Copy", "Intro", "Thank you for supporting our mission and helping create more opportunities for the community.", ""),
                ("Copy", "Footer emphasis", "Stronger together. / Shared goals. Greater impact.", ""),
                ("Image", "Title decoration", "assets/decor/partners-handshake-transparent.png", ""),
                ("Image", "Honeycomb graphic", "assets/partners/partner-honeycomb.png", ""),
            ]
            + [
                ("Image", f"Partner logo {i} — {name}", path, "")
                for i, (name, path) in enumerate(
                    [
                        ("Otago Polytechnic", "assets/partners/logos/otago-polytechnic.png"),
                        ("IIBA", "assets/partners/logos/iiba.png"),
                        ("AUT", "assets/partners/logos/aut.png"),
                        ("redvespa", "assets/partners/logos/redvespa.png"),
                        ("Victoria University of Wellington", "assets/partners/logos/victoria-wellington.png"),
                        ("nextwork", "assets/partners/logos/nextwork.png"),
                        ("Dev Academy Aotearoa", "assets/partners/logos/dev-academy.png"),
                        ("Potentia", "assets/partners/logos/potentia.png"),
                        ("Summer of Tech", "assets/partners/logos/summer-of-tech.png"),
                        ("CITANZ", "assets/partners/logos/citanz.png"),
                        ("Bridging the Gap", "assets/partners/logos/bridging-the-gap.png"),
                        ("Kiwibank", "assets/partners/logos/kiwibank.png"),
                        ("Inside Business Analysis", "assets/partners/logos/inside-ba.png"),
                        ("Powrsuit", "assets/partners/logos/powsuit.png"),
                        ("ProductTank", "assets/partners/logos/product-tank.png"),
                        ("ba&beyond", "assets/partners/logos/ba-beyond.png"),
                        ("Timmy", "assets/partners/logos/timmy.png"),
                        ("INSIDE on purpose", "assets/partners/logos/inside-on-purpose.png"),
                        ("Hudson", "assets/partners/logos/hudson.png"),
                    ],
                    1,
                )
            ],
        ),
        "07 Thank You Volunteers": (
            "07 · Thank You — Volunteers",
            None,
            [
                ("Copy", "Title", "Our Volunteers", ""),
                ("Copy", "Intro", "Thank you for your time, energy and care behind the scenes. You helped make this community possible.", ""),
                ("Copy", "Card 1 caption", "Planning with care", ""),
                ("Copy", "Card 2 caption", "Helping with heart", ""),
                ("Copy", "Footer emphasis", "Your dedication makes the difference. / Grateful for all you do.", ""),
                ("Image", "Title decoration", "assets/decor/volunteers-heart-hand-transparent.png", ""),
                ("Image", "Main photo", "assets/thanks/volunteers-1.jpg", ""),
                ("Image", "Photo 2", "assets/thanks/volunteers-2.jpg", ""),
                ("Image", "Photo 3", "assets/thanks/volunteers-3.jpg", ""),
            ],
        ),
        "08 Whats Next": (
            "08 · What's Next",
            None,
            [
                ("Copy", "Title", "What's Next", ""),
                ("Copy", "Subtitle", "Here's what we hope to keep building together.", ""),
                ("Copy", "Card 1 title", "More practical career support", ""),
                ("Copy", "Card 1 detail", "Guidance, resources and support for BA careers.", ""),
                ("Copy", "Card 2 title", "More real-world BA stories", ""),
                ("Copy", "Card 2 detail", "Stories and insights from real BA journeys.", ""),
                ("Copy", "Card 3 title", "More opportunities to connect", ""),
                ("Copy", "Card 3 detail", "More ways to learn, meet and grow together.", ""),
                ("Copy", "Bottom photo caption", "Still growing, together. ♡", ""),
                ("Image", "Card icon 1", "assets/whats-next/icon-career-support-transparent.png", ""),
                ("Image", "Card icon 2", "assets/whats-next/icon-ba-stories-transparent.png", ""),
                ("Image", "Card icon 3", "assets/whats-next/icon-connect-transparent.png", ""),
                ("Image", "Bottom group photo", "assets/whats-next/cutted.png", ""),
            ],
        ),
        "09 Your Voice": (
            "09 · Your Voice",
            "Form submission requires Formspree endpoint before launch.",
            [
                ("Copy", "Title", "Your Voice", ""),
                ("Copy", "Subtitle", "is part of our story.", ""),
                ("Copy", "Question", "What has BA Career meant to you?", ""),
                ("Copy", "Prompt", "Share a memory, a message, or an idea for what you would like to see next.", ""),
                ("Copy", "Input placeholder", "Write your message...", ""),
                ("Copy", "Character limit", "500", ""),
                ("Copy", "Submit button", "Share your message", ""),
                ("Copy", "Success title", "Message shared!", ""),
                ("Copy", "Success message", "Thank you for adding your voice to our story.", ""),
                ("Copy", "Post-submit note", "Your message will help us celebrate where we have been and shape what comes next.", ""),
                ("Image", "Title decoration", "assets/decor/speakers-heart-transparent.png", ""),
                ("Image", "Envelope illustration", "assets/feedback/envelope-heart.png", ""),
                ("Config", "Formspree endpoint", "(empty — not connected yet)", ""),
            ],
        ),
        "10 Closing": (
            "10 · Closing",
            "Closing is a full artwork image. Text changes require a new design file unless layout is rebuilt.",
            [
                ("Image", "Full-page artwork", "assets/ending/ChatGPT Image Jul 13, 2026, 06_41_08 PM.png", ""),
                ("Copy", "Headline (in artwork / config)", "Here's to", ""),
                ("Copy", "Headline script (in artwork / config)", "3,000 Members", ""),
                ("Copy", "Subline (in artwork / config)", "and many more to come!", ""),
                ("Copy", "CTA (in artwork / config)", "Stay connected with us", ""),
                ("Copy", "Tagline (in artwork / config)", "See you at the next meetup!", ""),
                ("Copy", "Footer script (in artwork / config)", "Grow with us.\nWe grow together.", ""),
                ("Link", "Meetup", "https://www.meetup.com/ba-career-meetup-group/", ""),
                ("Link", "LinkedIn", "https://www.linkedin.com/company/ba-career", ""),
                ("Link", "YouTube", "https://www.youtube.com/@bacareermeetup", ""),
                ("Link", "Website", "https://analysis.nz/", ""),
            ],
        ),
    }

    for sheet_name, (title, note, rows) in pages.items():
        ws = wb.create_sheet(sheet_name)
        write_sheet(ws, title, rows, note)

    # --- Global & Sharing ---
    gs = wb.create_sheet("Global and Sharing")
    write_sheet(
        gs,
        "Global Brand & LinkedIn Sharing",
        [
            ("Copy", "Brand name", "BA CAREER MEETUP", ""),
            ("Copy", "Tagline", "GROW WITH US", ""),
            ("Image", "Logo (opening & top bar)", "assets/icons/Icon black.png", ""),
            ("Copy", "Browser title", "BA Career Meetup — 3,000 Members Celebration", ""),
            ("Copy", "LinkedIn share title (og:title)", "BA Career Meetup — 3,000 Members Celebration", ""),
            ("Copy", "LinkedIn share description (og:description)", "A journey from our first meetup to 3,000 members. Thank you to everyone who made this community possible.", ""),
            ("Image", "LinkedIn preview image (og:image)", "assets/og/og-image.jpg (1200×630 — pending)", ""),
        ],
        "Inner pages show BA CAREER MEETUP + logo in the top bar (except Opening).",
    )

    # --- Consistency Rules ---
    cr = wb.create_sheet("Consistency Rules")
    cr["A1"] = "Consistency Rules — Anchor Values"
    cr["A1"].font = TITLE_FONT
    cr.merge_cells("A1:D1")
    cr["A2"] = "When updating any value below, check all listed page locations."
    cr["A2"].font = NOTE_FONT
    cr.merge_cells("A2:D2")

    cr_headers = ["Anchor", "Current Value", "Appears On", "Update"]
    for i, h in enumerate(cr_headers, 1):
        cr.cell(row=4, column=i, value=h)
    style_header_row(cr, 4, 4)

    rules = [
        ("Member milestone", "3,000 / 3,000+", "01 Opening, 02 Journey subtitle, 02 2026 node, 03 Members, 10 Closing", ""),
        ("2025 milestone", "2,000 members", "02 Journey 2025 node", ""),
        ("Community start year", "2023", "02 Journey", ""),
        ("Highlight year", "2026", "02 Journey", ""),
        ("Events hosted", "40+", "03 By the Numbers", ""),
        ("Guest speakers", "60+", "03 By the Numbers", ""),
        ("Volunteers", "22+", "03 By the Numbers", ""),
        ("Member countries", "35+", "03 By the Numbers", ""),
        ("Event rating", "4.8 / 5 (341 feedbacks)", "03 By the Numbers", ""),
        ("Brand name", "BA Career / BA CAREER MEETUP", "Top bar, share meta", ""),
        ("External links", "Meetup, LinkedIn, YouTube, Website", "10 Closing artwork, config", ""),
    ]
    for idx, row in enumerate(rules, 5):
        for col, val in enumerate(row, 1):
            cr.cell(row=idx, column=col, value=val)
    style_data_rows(cr, 5, 5 + len(rules) - 1, 4, 4)
    cr.column_dimensions["A"].width = 22
    cr.column_dimensions["B"].width = 28
    cr.column_dimensions["C"].width = 48
    cr.column_dimensions["D"].width = 40
    cr.freeze_panes = "A5"

    # --- Assets Pending ---
    ap = wb.create_sheet("Assets Pending")
    write_sheet(
        ap,
        "Assets Still Needed",
        [
            ("Image", "01 Opening collages", "assets/opening/collage-1.jpg … collage-6.jpg", ""),
            ("Image", "04 Community Moments", "assets/moments/moment-1.jpg … moment-5.jpg", ""),
            ("Image", "05 Speakers photos", "assets/thanks/speakers-1.jpg … speakers-3.jpg", ""),
            ("Image", "07 Volunteers photos", "assets/thanks/volunteers-1.jpg … volunteers-3.jpg", ""),
            ("Image", "LinkedIn OG image", "assets/og/og-image.jpg (1200×630)", ""),
            ("Config", "Feedback form endpoint", "Formspree URL (before launch)", ""),
        ],
        "Priority list for ops / design team.",
    )

    return wb


if __name__ == "__main__":
    import os
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(root, OUTPUT)
    wb = build_workbook()
    wb.save(path)
    print(f"Saved: {path}")
